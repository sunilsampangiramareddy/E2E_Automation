import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class WriteExcelResults:
    def __init__(self, script_name, file_path=None):
        self.script_name = script_name
        if file_path is None:
            self.file_path = self.generate_file_path()
        else:
            self.file_path = file_path

        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def generate_file_path(self):
        # Get the current date and time
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # Combine script name and timestamp
        file_name = f"{self.script_name}_{current_time}.xlsx"
        # Define the directory to save the file
        directory = "C:\\Automation_Project\\E2E_Automation\\excel_test_results"
        # Ensure the directory exists
        if not os.path.exists(directory):
            os.makedirs(directory)
        # Return the full file path
        return os.path.join(directory, file_name)

    def generate_sheet_name(self):
        # Get the current date and time
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # Combine script name and timestamp
        return f"{self.script_name}_{current_time}"

    def write_data(self, data, sheet_name=None):
        if sheet_name is None:
            sheet_name = self.generate_sheet_name()
        self.sheet.title = sheet_name
        for row_index, row_data in enumerate(data, start=1):
            for col_index, value in enumerate(row_data, start=1):
                col_letter = get_column_letter(col_index)
                self.sheet[f"{col_letter}{row_index}"] = value
        self.save()

    def save(self):
        self.workbook.save(self.file_path)

    def append_data(self, data, sheet_name=None):
        if sheet_name is None:
            sheet_name = self.generate_sheet_name()
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.create_sheet(title=sheet_name)
        for row_data in data:
            self.sheet.append(row_data)
        self.save()
